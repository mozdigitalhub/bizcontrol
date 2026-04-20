from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
import re
import unicodedata

from django.db import transaction
from django.db.models import Sum
from django.utils.text import slugify

from catalog.models import Category, Product
from inventory.models import StockMovement
from inventory.services import record_movement

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


REQUIRED_HEADERS = {
    "descricao_do_item": "name",
    "un": "unit",
    "preco_de_compra": "cost",
    "preco_de_venda": "sale",
    "entrada": "qty",
}


@dataclass
class ImportResult:
    categories_created: int = 0
    products_created: int = 0
    products_updated: int = 0
    rows_processed: int = 0
    rows_failed: int = 0
    failed_products_count: int = 0
    failed_products: list = field(default_factory=list)
    errors: list = field(default_factory=list)


class ExcelImportService:
    def __init__(self, *, business, user=None):
        self.business = business
        self.user = user

    def parse_workbook(self, file_obj):
        if not openpyxl:
            raise ValueError("Biblioteca openpyxl nao instalada.")
        return openpyxl.load_workbook(file_obj, data_only=True)

    def import_sheet(self, sheet, result: ImportResult):
        sheet_name = (sheet.title or "").strip()
        if not sheet_name:
            return

        category = Category.objects.filter(
            business=self.business,
            name__iexact=sheet_name,
        ).first()
        if not category:
            category = Category.objects.create(business=self.business, name=sheet_name)
            result.categories_created += 1

        header_map, missing_headers = self._resolve_headers(sheet)
        if not header_map:
            self._add_error(
                result=result,
                sheet=sheet_name,
                row=1,
                product="-",
                error=(
                    "Cabecalho invalido ou incompleto. "
                    f"Campos em falta: {', '.join(missing_headers) if missing_headers else 'desconhecido'}."
                ),
            )
            return

        products_qs = Product.objects.filter(business=self.business)
        existing_products = {self._normalize_name(p.name): p for p in products_qs}
        seen_rows = {}

        product_ids = [p.id for p in existing_products.values()]
        stock_map = self._get_stock_map(product_ids)

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if self._row_is_empty(row):
                continue
            result.rows_processed += 1
            product_name = self._safe_str(row[header_map["name"]]) if "name" in header_map else ""
            try:
                data = self._extract_row(row, header_map)
                if not data["name"]:
                    raise ValueError("Descricao do item obrigatoria.")
                if not data["unit"]:
                    raise ValueError("Unidade obrigatoria.")
                if data["qty"] < 0:
                    raise ValueError("Entrada nao pode ser negativa.")

                normalized_name = self._normalize_name(data["name"])
                if normalized_name in seen_rows:
                    first_row = seen_rows[normalized_name]
                    raise ValueError(
                        f"Produto duplicado no ficheiro. Ja informado na linha {first_row}."
                    )
                seen_rows[normalized_name] = row_idx

                product = existing_products.get(normalized_name)
                if product:
                    updated = self._update_product(product, data)
                    if updated:
                        result.products_updated += 1
                else:
                    product = self._create_product(category, data)
                    existing_products[normalized_name] = product
                    result.products_created += 1

                current_stock = stock_map.get(product.id, 0)
                target_stock = data["qty"]
                delta = target_stock - current_stock
                if delta != 0:
                    movement_type = (
                        StockMovement.MOVEMENT_IN
                        if delta > 0
                        else StockMovement.MOVEMENT_OUT
                    )
                    record_movement(
                        business=self.business,
                        product=product,
                        movement_type=movement_type,
                        quantity=abs(delta),
                        created_by=self.user,
                        reference_type="initial_import",
                        notes="Importacao inicial",
                    )
                    stock_map[product.id] = target_stock
            except Exception as exc:
                self._add_error(
                    result=result,
                    sheet=sheet_name,
                    row=row_idx,
                    product=product_name or "-",
                    error=str(exc),
                )

    def import_workbook(self, file_obj):
        workbook = self.parse_workbook(file_obj)
        result = ImportResult()
        for sheet in workbook.worksheets:
            with transaction.atomic():
                self.import_sheet(sheet, result)
        return result

    def _add_error(self, *, result, sheet, row, error, product="-"):
        result.rows_failed += 1
        result.errors.append(
            {
                "sheet": sheet,
                "row": row,
                "product": product,
                "error": error,
            }
        )
        if product and product != "-":
            result.failed_products.append(product)
            result.failed_products = list(dict.fromkeys(result.failed_products))
        result.failed_products_count = len(result.failed_products)

    def _resolve_headers(self, sheet):
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return None, list(REQUIRED_HEADERS.keys())

        mapping = {}
        for idx, header in enumerate(header_row):
            if header is None:
                continue
            key = self._normalize_header(str(header))
            if key in REQUIRED_HEADERS:
                mapping[REQUIRED_HEADERS[key]] = idx

        reverse_required = {v: k for k, v in REQUIRED_HEADERS.items()}
        missing = [
            reverse_required[field]
            for field in REQUIRED_HEADERS.values()
            if field not in mapping
        ]
        if missing:
            return None, missing
        return mapping, []

    def _extract_row(self, row, mapping):
        name = self._safe_str(row[mapping["name"]])
        unit = self._safe_str(row[mapping["unit"]])
        cost = self._to_decimal(row[mapping["cost"]])
        sale = self._to_decimal(row[mapping["sale"]])
        qty = self._to_int(row[mapping["qty"]])
        if cost is None:
            cost = Decimal("0")
        if sale is None:
            sale = Decimal("0")
        if qty is None:
            qty = 0
        return {
            "name": name,
            "unit": unit,
            "cost": cost,
            "sale": sale,
            "qty": qty,
        }

    def _create_product(self, category, data):
        sku = self._generate_sku(data["name"])
        return Product.objects.create(
            business=self.business,
            name=data["name"],
            sku=sku,
            category=category,
            unit_of_measure=self._normalize_unit(data["unit"]),
            cost_price=data["cost"] or 0,
            sale_price=data["sale"],
            created_by=self.user,
            updated_by=self.user,
        )

    def _update_product(self, product, data):
        changed = False
        unit = self._normalize_unit(data["unit"])
        if unit and product.unit_of_measure != unit:
            product.unit_of_measure = unit
            changed = True
        if data["sale"] is not None and product.sale_price != data["sale"]:
            product.sale_price = data["sale"]
            changed = True
        if data["cost"] is not None and product.cost_price != data["cost"]:
            product.cost_price = data["cost"]
            changed = True
        if changed:
            product.updated_by = self.user
            product.save(
                update_fields=[
                    "unit_of_measure",
                    "sale_price",
                    "cost_price",
                    "updated_by",
                ]
            )
        return changed

    def _get_stock_map(self, product_ids):
        if not product_ids:
            return {}
        totals = (
            StockMovement.objects.filter(
                business=self.business,
                product_id__in=product_ids,
            )
            .values("product_id", "movement_type")
            .annotate(total=Sum("quantity"))
        )
        summary = {}
        for row in totals:
            summary.setdefault(row["product_id"], {})[row["movement_type"]] = row["total"] or 0

        stock_map = {}
        for product_id, data in summary.items():
            incoming = data.get(StockMovement.MOVEMENT_IN, 0)
            outgoing = data.get(StockMovement.MOVEMENT_OUT, 0)
            adjust = data.get(StockMovement.MOVEMENT_ADJUST, 0)
            stock_map[product_id] = int(incoming) - int(outgoing) + int(adjust)
        return stock_map

    def _generate_sku(self, name):
        base = slugify(name)[:16] or "item"
        suffix = hashlib.md5(name.encode("utf-8")).hexdigest()[:4]
        sku = f"{base}-{suffix}"
        counter = 1
        while Product.objects.filter(business=self.business, sku=sku).exists():
            sku = f"{base}-{suffix}{counter}"
            counter += 1
        return sku

    def _normalize_header(self, value):
        if not value:
            return ""
        value = unicodedata.normalize("NFKD", value)
        value = value.encode("ascii", "ignore").decode("ascii")
        return value.strip().lower().replace(".", "").replace(" ", "_")

    def _normalize_name(self, value):
        value = (value or "").strip().lower()
        value = unicodedata.normalize("NFKD", value)
        value = value.encode("ascii", "ignore").decode("ascii")
        value = re.sub(r"[^a-z0-9]+", " ", value)
        return " ".join(value.split())

    def _normalize_unit(self, value):
        value = (value or "").strip().lower()
        map_units = {
            "un": "un",
            "unid": "un",
            "unidade": "un",
            "kg": "kg",
            "kilo": "kg",
            "metro": "metro",
            "m": "metro",
            "saco": "saco",
            "litro": "litro",
            "l": "litro",
            "caixa": "caixa",
            "pacote": "pacote",
        }
        return map_units.get(value, "un")

    def _row_is_empty(self, row):
        return all(cell is None or str(cell).strip() == "" for cell in row)

    def _safe_str(self, value):
        return str(value).strip() if value is not None else ""

    def _to_decimal(self, value):
        if value is None or str(value).strip() == "":
            return None
        raw = str(value).strip().replace(" ", "")
        raw = raw.replace("MZN", "").replace("MT", "")
        if "," in raw and "." in raw:
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", ".")
        try:
            return Decimal(raw)
        except (InvalidOperation, ValueError):
            raise ValueError(f"Valor numerico invalido: {value}.")

    def _to_int(self, value):
        if value is None or str(value).strip() == "":
            return None
        try:
            dec = self._to_decimal(value)
            if dec is None:
                return None
            return int(dec.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        except (InvalidOperation, ValueError):
            raise ValueError(f"Quantidade invalida: {value}.")
