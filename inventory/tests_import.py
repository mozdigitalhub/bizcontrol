from io import BytesIO
from decimal import Decimal

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase

from catalog.models import Category, Product
from inventory.excel_import import ExcelImportService
from inventory.models import StockMovement
from tenants.models import Business, BusinessMembership


class StockExcelImportTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="importer", password="pass")
        self.business = Business.objects.create(name="Ferragem", slug="ferragem", business_type=Business.BUSINESS_HARDWARE)
        BusinessMembership.objects.create(
            business=self.business, user=self.user, role=BusinessMembership.ROLE_OWNER
        )

    def _build_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cimento"
        ws.append([
            "Descricao do Item",
            "Un.",
            "PRECO DE COMPRA",
            "PRECO DE VENDA",
            "ENTRADA",
        ])
        ws.append(["Cimento 32", "saco", "250,00", "450,00", "10"])
        ws.append(["Cimento 42", "saco", "", "500", "5"])
        return wb

    def test_import_creates_categories_and_products(self):
        wb = self._build_workbook()
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        service = ExcelImportService(business=self.business, user=self.user)
        result = service.import_workbook(buffer)
        self.assertEqual(Category.objects.filter(business=self.business).count(), 1)
        self.assertEqual(Product.objects.filter(business=self.business).count(), 2)
        self.assertEqual(result.products_created, 2)
        self.assertEqual(StockMovement.objects.filter(business=self.business).count(), 2)

    def test_import_updates_existing_product(self):
        category = Category.objects.create(business=self.business, name="Cimento")
        product = Product.objects.create(
            business=self.business,
            category=category,
            name="Cimento 32",
            sale_price=Decimal("400.00"),
            cost_price=Decimal("200.00"),
        )
        wb = self._build_workbook()
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        service = ExcelImportService(business=self.business, user=self.user)
        result = service.import_workbook(buffer)
        product.refresh_from_db()
        self.assertEqual(product.sale_price, Decimal("450.00"))
        self.assertEqual(result.products_updated, 1)

    def test_missing_required_fields_reported(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ferramentas"
        ws.append([
            "Descricao do Item",
            "Un.",
            "PRECO DE COMPRA",
            "PRECO DE VENDA",
            "ENTRADA",
        ])
        ws.append(["", "un", "", "", ""])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        service = ExcelImportService(business=self.business, user=self.user)
        result = service.import_workbook(buffer)
        self.assertEqual(result.rows_failed, 1)
        self.assertTrue(result.errors)
        self.assertEqual(result.failed_products_count, 0)

    def test_duplicate_product_in_same_file_is_reported_with_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ferramentas"
        ws.append(
            [
                "Descricao do Item",
                "Un.",
                "PRECO DE COMPRA",
                "PRECO DE VENDA",
                "ENTRADA",
            ]
        )
        ws.append(["Martelo 500g", "un", "100", "180", "3"])
        ws.append(["Martelo 500g", "un", "100", "180", "4"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        service = ExcelImportService(business=self.business, user=self.user)
        result = service.import_workbook(buffer)

        self.assertEqual(Product.objects.filter(business=self.business).count(), 1)
        self.assertEqual(result.rows_failed, 1)
        self.assertEqual(result.failed_products_count, 1)
        self.assertIn("Martelo 500g", result.failed_products)
        self.assertIn("linha 2", result.errors[0]["error"].lower())

    def test_import_does_not_duplicate_existing_product_with_name_variation(self):
        category = Category.objects.create(business=self.business, name="Cimento")
        Product.objects.create(
            business=self.business,
            category=category,
            name="Cimento Nacional 32",
            sale_price=Decimal("420.00"),
            cost_price=Decimal("210.00"),
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cimento"
        ws.append(
            [
                "Descricao do Item",
                "Un.",
                "PRECO DE COMPRA",
                "PRECO DE VENDA",
                "ENTRADA",
            ]
        )
        ws.append(["Cimento  Nacional   32", "saco", "260", "460", "8"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        service = ExcelImportService(business=self.business, user=self.user)
        result = service.import_workbook(buffer)

        self.assertEqual(Product.objects.filter(business=self.business).count(), 1)
        self.assertEqual(result.products_created, 0)
        self.assertEqual(result.products_updated, 1)

    def test_import_uses_zero_defaults_for_blank_qty_and_prices(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Acessorios"
        ws.append(
            [
                "Descricao do Item",
                "Un.",
                "PRECO DE COMPRA",
                "PRECO DE VENDA",
                "ENTRADA",
            ]
        )
        ws.append(["Parafuso 8mm", "un", "", "", ""])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        service = ExcelImportService(business=self.business, user=self.user)
        result = service.import_workbook(buffer)

        product = Product.objects.get(business=self.business, name="Parafuso 8mm")
        self.assertEqual(product.cost_price, Decimal("0"))
        self.assertEqual(product.sale_price, Decimal("0"))
        self.assertEqual(result.rows_failed, 0)
        # Quantidade vazia vira 0 e nao gera movimento de stock.
        self.assertFalse(
            StockMovement.objects.filter(
                business=self.business,
                product=product,
                reference_type="initial_import",
            ).exists()
        )
